"""
畅点餐 - Excel 导出路由
支持订单导出、菜品销售报表导出
使用 openpyxl 生成专业格式的 Excel 文件
"""
import io
import os
from datetime import datetime, date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import Order, OrderItem, Dish, Store, Table

router = APIRouter(prefix="/export", tags=["导出"])

# ── 样式常量 ────────────────────────────────────────────────────

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# ── 辅助函数 ────────────────────────────────────────────────────

def _setup_header(ws, headers, col_widths=None):
    """设置表头样式"""
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    # 设置列宽
    if col_widths:
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    else:
        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16
    # 冻结首行
    ws.freeze_panes = "A2"


def _write_data_row(ws, row_idx, values, is_alt=False):
    """写入一行数据并应用样式"""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        # 数字类型右对齐
        if isinstance(val, (int, float)):
            cell.alignment = NUMBER_ALIGNMENT
        else:
            cell.alignment = DATA_ALIGNMENT
        if is_alt:
            cell.fill = ALT_FILL


# ── 状态映射 ────────────────────────────────────────────────────

ORDER_TYPE_MAP = {
    "dine_in": "堂食",
    "takeaway": "自提",
    "delivery": "外卖",
}

ORDER_STATUS_MAP = {
    "pending": "待支付",
    "paid": "已支付",
    "preparing": "制作中",
    "ready": "待上菜",
    "served": "已上菜",
    "completed": "已完成",
    "cancelled": "已取消",
    "refunding": "退款中",
    "refunded": "已退款",
}

PAY_STATUS_MAP = {
    "unpaid": "未支付",
    "paid": "已支付",
    "refunded": "已退款",
}

PAY_TYPE_MAP = {
    "wechat": "微信支付",
    "balance": "余额支付",
    "": "未选择",
}


def _get_store_name(db: Session, store_id: int) -> str:
    store = db.query(Store).filter(Store.id == store_id).first()
    return store.name if store else f"门店{store_id}"


def _get_table_no(db: Session, table_id: int) -> str:
    table = db.query(Table).filter(Table.id == table_id).first()
    return table.table_no if table else ""


# ── 订单导出 ────────────────────────────────────────────────────

@router.get("/orders")
def export_orders(
    store_id: int = Query(..., description="门店ID"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="订单状态筛选"),
    order_type: Optional[str] = Query(None, description="订单类型筛选 (dine_in/takeaway/delivery)"),
    db: Session = Depends(get_db),
):
    """
    导出订单列表到 Excel
    列: 订单号, 下单时间, 订单类型, 订单状态, 支付状态, 支付方式, 桌号, 配送地址, 商品明细, 原价, 优惠, 实付金额, 配送费
    """
    # 查询订单
    query = db.query(Order).filter(Order.store_id == store_id)

    if start_date:
        try:
            dt_start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Order.created_at >= dt_start)
        except ValueError:
            raise HTTPException(status_code=400, detail="start_date 格式应为 YYYY-MM-DD")
    if end_date:
        try:
            dt_end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(Order.created_at <= dt_end)
        except ValueError:
            raise HTTPException(status_code=400, detail="end_date 格式应为 YYYY-MM-DD")
    if status:
        query = query.filter(Order.status == status)
    if order_type:
        query = query.filter(Order.order_type == order_type)

    orders = query.order_by(Order.created_at.desc()).all()

    if not orders:
        raise HTTPException(status_code=404, detail="该条件下没有订单数据")

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    store_name = _get_store_name(db, store_id)
    ws.title = "订单明细"

    # 表头
    headers = [
        "订单号", "下单时间", "订单类型", "订单状态", "支付状态",
        "支付方式", "桌号", "配送地址", "商品明细", "原价",
        "优惠", "实付金额", "配送费",
    ]
    col_widths = [22, 20, 10, 10, 10, 12, 10, 28, 40, 12, 12, 12, 10]
    _setup_header(ws, headers, col_widths)
    ws.row_dimensions[1].height = 28

    # 写入数据
    for row_idx, order in enumerate(orders, 2):
        # 获取桌号
        table_no = ""
        if order.table_id:
            table_no = _get_table_no(db, order.table_id)

        # 获取商品明细
        items = db.query(OrderItem).filter(OrderItem.order_id == order.id).all()
        items_detail = "\n".join(
            f"{item.dish_name} x{item.quantity}"
            for item in items
        )

        # 配送地址处理
        delivery_addr = order.delivery_address or ""
        if delivery_addr and order.delivery_name:
            delivery_addr = f"{order.delivery_name} {order.delivery_phone}\n{delivery_addr}"

        values = [
            order.order_no,
            order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else "",
            ORDER_TYPE_MAP.get(order.order_type, order.order_type),
            ORDER_STATUS_MAP.get(order.status, order.status),
            PAY_STATUS_MAP.get(order.pay_status, order.pay_status),
            PAY_TYPE_MAP.get(order.pay_type, order.pay_type),
            table_no,
            delivery_addr,
            items_detail,
            round(order.total_amount or 0, 2),
            round(order.discount_amount or 0, 2),
            round(order.pay_amount or 0, 2),
            round(order.delivery_fee or 0, 2),
        ]
        _write_data_row(ws, row_idx, values, is_alt=(row_idx % 2 == 0))
        ws.row_dimensions[row_idx].height = max(30, 16 * len(items))

    # 汇总行
    summary_row = len(orders) + 2
    total_original = sum(o.total_amount or 0 for o in orders)
    total_discount = sum(o.discount_amount or 0 for o in orders)
    total_pay = sum(o.pay_amount or 0 for o in orders)
    total_delivery = sum(o.delivery_fee or 0 for o in orders)

    ws.cell(row=summary_row, column=1, value="汇总").font = Font(
        name="微软雅黑", size=11, bold=True, color="FFFFFF"
    )
    ws.cell(row=summary_row, column=1).fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    for col in range(1, len(headers) + 1):
        ws.cell(row=summary_row, column=col).fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        ws.cell(row=summary_row, column=col).border = THIN_BORDER

    ws.cell(row=summary_row, column=10, value=round(total_original, 2)).font = Font(
        name="微软雅黑", size=11, bold=True, color="FFFFFF"
    )
    ws.cell(row=summary_row, column=11, value=round(total_discount, 2)).font = Font(
        name="微软雅黑", size=11, bold=True, color="FFFFFF"
    )
    ws.cell(row=summary_row, column=12, value=round(total_pay, 2)).font = Font(
        name="微软雅黑", size=11, bold=True, color="FFFFFF"
    )
    ws.cell(row=summary_row, column=13, value=round(total_delivery, 2)).font = Font(
        name="微软雅黑", size=11, bold=True, color="FFFFFF"
    )

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 构建文件名
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    date_range = ""
    if start_date and end_date:
        date_range = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_range = f"_from_{start_date}"
    elif end_date:
        date_range = f"_to_{end_date}"

    filename = f"orders_{store_name}{date_range}_{date_str}.xlsx"

    return FileResponse(
        path=None,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=output.read(),
    )


# ── 菜品销售报表 ────────────────────────────────────────────────

@router.get("/dishes")
def export_dishes(
    store_id: int = Query(..., description="门店ID"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    """
    导出菜品销售报表到 Excel
    列: 菜品名称, 分类, 单价, 销量, 销售额, 库存状态
    按销量降序排列
    """
    # 查询菜品及其分类信息
    dishes = (
        db.query(Dish)
        .filter(Dish.store_id == store_id)
        .order_by(Dish.sales_count.desc())
        .all()
    )

    if not dishes:
        raise HTTPException(status_code=404, detail="该门店没有菜品数据")

    # 如果有日期范围，计算该时间段内的实际销量
    dish_sales = {}
    if start_date or end_date:
        # 构建日期条件
        date_filters = [Order.store_id == store_id]
        date_filters.append(Order.status.notin_(["cancelled", "refunded"]))

        if start_date:
            try:
                dt_start = datetime.strptime(start_date, "%Y-%m-%d")
                date_filters.append(Order.created_at >= dt_start)
            except ValueError:
                raise HTTPException(status_code=400, detail="start_date 格式应为 YYYY-MM-DD")
        if end_date:
            try:
                dt_end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                date_filters.append(Order.created_at <= dt_end)
            except ValueError:
                raise HTTPException(status_code=400, detail="end_date 格式应为 YYYY-MM-DD")

        # 聚合各菜品的实际销量和销售额
        sales_data = (
            db.query(
                OrderItem.dish_id,
                OrderItem.dish_name,
                func.sum(OrderItem.quantity).label("total_qty"),
                func.sum(OrderItem.subtotal).label("total_revenue"),
            )
            .join(Order, OrderItem.order_id == Order.id)
            .filter(and_(*date_filters))
            .group_by(OrderItem.dish_id, OrderItem.dish_name)
            .all()
        )

        for s in sales_data:
            dish_sales[s.dish_id] = {
                "quantity": int(s.total_qty or 0),
                "revenue": float(s.total_revenue or 0),
            }

    wb = Workbook()
    ws = wb.active
    store_name = _get_store_name(db, store_id)
    ws.title = "菜品销售报表"

    # 表头
    headers = ["菜品名称", "分类", "单价", "销量", "销售额", "库存状态"]
    col_widths = [24, 14, 12, 10, 14, 12]
    _setup_header(ws, headers, col_widths)
    ws.row_dimensions[1].height = 28

    # 写入数据
    total_revenue = 0.0
    total_sales = 0

    for row_idx, dish in enumerate(dishes, 2):
        # 获取分类名
        category_name = ""
        if dish.category:
            category_name = dish.category.name

        # 销量和销售额
        if dish.dish_id in dish_sales:
            sales_qty = dish_sales[dish.dish_id]["quantity"]
            sales_revenue = dish_sales[dish.dish_id]["revenue"]
        else:
            if start_date or end_date:
                sales_qty = 0
                sales_revenue = 0.0
            else:
                sales_qty = dish.sales_count or 0
                sales_revenue = (dish.sales_count or 0) * (dish.price or 0)

        # 库存状态
        if dish.stock <= 0:
            stock_status = "售罄"
        elif dish.stock <= 10:
            stock_status = "紧张"
        elif dish.stock <= 50:
            stock_status = "正常"
        else:
            stock_status = "充足"

        total_revenue += sales_revenue
        total_sales += sales_qty

        values = [
            dish.name,
            category_name,
            round(dish.price or 0, 2),
            sales_qty,
            round(sales_revenue, 2),
            stock_status,
        ]
        _write_data_row(ws, row_idx, values, is_alt=(row_idx % 2 == 0))

        # 根据库存状态设置颜色
        status_cell = ws.cell(row=row_idx, column=6)
        if stock_status == "售罄":
            status_cell.font = Font(name="微软雅黑", size=10, color="C00000")
        elif stock_status == "紧张":
            status_cell.font = Font(name="微软雅黑", size=10, color="ED7D31")

    # 汇总行
    summary_row = len(dishes) + 2
    ws.cell(row=summary_row, column=1, value=f"汇总 (共 {len(dishes)} 个菜品)").font = Font(
        name="微软雅黑", size=11, bold=True, color="FFFFFF"
    )
    for col in range(1, len(headers) + 1):
        ws.cell(row=summary_row, column=col).fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        ws.cell(row=summary_row, column=col).border = THIN_BORDER

    ws.cell(row=summary_row, column=4, value=total_sales).font = Font(
        name="微软雅黑", size=11, bold=True, color="FFFFFF"
    )
    ws.cell(row=summary_row, column=5, value=round(total_revenue, 2)).font = Font(
        name="微软雅黑", size=11, bold=True, color="FFFFFF"
    )

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    date_range = ""
    if start_date and end_date:
        date_range = f"_{start_date}_to_{end_date}"
    filename = f"dishes_{store_name}{date_range}_{date_str}.xlsx"

    return FileResponse(
        path=None,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=output.read(),
    )
